import datetime
from decimal import Decimal
import openpyxl
from models import SheetEntry, EntriesReader

MAX_COL = 8
EXCEL_OFFSET = 1
COL = {"DATE": "A",
       "HISTORY": "D",
       "VALUE": "F",
       "DESC": "C",
       "BANKID": "G"}

PAYMENT = "PAGAMENTOS"
DEPOSIT = "RECEBIMENTOS"

class SheetReader(EntriesReader):
    def __init__(self, sheet_path):
        self.wb = openpyxl.load_workbook(sheet_path)

    def all_entries(self):
        entries = []
        try:
            entries += self.get_entries(self.wb.get_sheet_by_name(PAYMENT), True)
            entries += self.get_entries(self.wb.get_sheet_by_name(DEPOSIT), False)
        except Exception:
            print("couldn't separate deposits from payments")
            entries += self.get_entries(self.wb.active)

        return entries

    def get_entries(self, ws, is_payment):
        entries = []
        start = self.find_first_row()
        for row in range(start, ws.max_row + EXCEL_OFFSET):
            value = ws[f"{COL['VALUE']}{row}"].value
            entry = SheetEntry(
                date=ws[f"{COL['DATE']}{row}"].value.strftime("%d/%m/%Y"),
                history=ws[f"{COL['HISTORY']}{row}"].value,
                value = Decimal(str(value)).__neg__() if is_payment else Decimal(str(value)),
                description=ws[f"{COL['DESC']}{row}"].value,
                bank_id=ws[f"{COL['BANKID']}{row}"].value,)
            entries.append(entry)
        return entries

    def find_first_row(self):
        ws = self.wb.active
        for n in range(EXCEL_OFFSET, ws.max_row + EXCEL_OFFSET):
            if type(ws[f"A{n}"].value) == datetime.datetime:
                return n
        raise Exception(
            "date not found, check if the sheet is in the correct format.")
